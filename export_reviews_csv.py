#!/usr/bin/env python3
"""
Script to export review details from Excel file to CSV format.
Extracts reviewer name, rating, date, review text, and helpful votes from all product sheets.
"""

import csv
import os
from openpyxl import load_workbook

def export_reviews_to_csv(excel_file, csv_file):
    """Export review data from Excel file to CSV"""

    # Load the Excel workbook
    wb = load_workbook(excel_file)

    # Prepare CSV data
    csv_data = []

    # Process each product sheet (skip Summary sheet)
    for sheet_name in wb.sheetnames:
        if sheet_name == "Summary":
            continue

        ws = wb[sheet_name]

        # Get product title from the first row
        product_title = ws['B1'].value if ws['B1'].value else "Unknown Product"

        # Find the review headers row (should contain "Reviewer Name")
        review_start_row = None
        for row_idx in range(1, ws.max_row + 1):
            row_values = [cell.value for cell in ws[row_idx]]
            if "Reviewer Name" in str(row_values):
                review_start_row = row_idx + 1  # Reviews start after headers
                break

        if review_start_row:
            # Extract reviews
            for row_idx in range(review_start_row, ws.max_row + 1):
                row = [cell.value for cell in ws[row_idx]]

                # Skip empty rows
                if not any(row):
                    continue

                # Extract review data
                reviewer_name = row[0] if len(row) > 0 else ""
                rating = row[1] if len(row) > 1 else ""
                date = row[2] if len(row) > 2 else ""
                review_text = row[3] if len(row) > 3 else ""
                helpful_votes = row[4] if len(row) > 4 else ""

                # Add to CSV data
                csv_data.append({
                    'Product': product_title,
                    'Reviewer Name': reviewer_name,
                    'Rating': rating,
                    'Date': date,
                    'Review Text': review_text,
                    'Helpful Votes': helpful_votes
                })

    # Write to CSV
    if csv_data:
        with open(csv_file, 'w', newline='', encoding='utf-8') as f:
            if csv_data:
                fieldnames = ['Product', 'Reviewer Name', 'Rating', 'Date', 'Review Text', 'Helpful Votes']
                writer = csv.DictWriter(f, fieldnames=fieldnames)
                writer.writeheader()
                writer.writerows(csv_data)

        print(f"✅ Exported {len(csv_data)} reviews to {csv_file}")
        return True
    else:
        print("❌ No review data found to export")
        return False

if __name__ == "__main__":
    # Use the existing Excel file
    excel_file = "exports/amazon_reviews_butter_20250916_094935.xlsx"
    csv_file = "exports/amazon_reviews_butter_20250916_094935.csv"

    if os.path.exists(excel_file):
        export_reviews_to_csv(excel_file, csv_file)
    else:
        print(f"❌ Excel file not found: {excel_file}")
