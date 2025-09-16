import asyncio
import random
import logging
import os
import tempfile
import json
from datetime import datetime
from flask import Flask, jsonify, request, send_file
from flask_cors import CORS
from bs4 import BeautifulSoup
from playwright.async_api import async_playwright
import csv
from openpyxl import Workbook

# ---------------- Logging setup ----------------
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(levelname)s - %(message)s"
)
logger = logging.getLogger(__name__)

# ---------------- User-Agent pool ----------------
USER_AGENTS = [
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
    "(KHTML, like Gecko) Chrome/123.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 "
    "(KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36",
    "Mozilla/5.0 (X11; Ubuntu; Linux x86_64; rv:123.0) "
    "Gecko/20100101 Firefox/123.0",
]

# ---------------- Authentication configuration ----------------
AUTO_AUTH_CONFIG = {
    "enabled": True,  # Enable by default since credentials are configured
    "credentials": {
        "email": "",
        "password": ""
    },
    "session_file": "amazon_session.json",
    "persistent_session": True
}

# ---------------- Timeout configuration ----------------
DEFAULT_TIMEOUT = 60000  # 60 seconds for most operations
NAVIGATION_TIMEOUT = 60000  # 60 seconds for navigation
SELECTOR_TIMEOUT = 15000  # 15 seconds for waiting for selectors

def create_app() -> Flask:
    app = Flask(__name__)

    # ---------------- Helper: Generate Excel file from review data ----------------
    def generate_excel_file(search_term, products):
        """Generate Excel file with review data organized by product"""

        # Create a new workbook and select the active worksheet
        wb = Workbook()
        ws = wb.active
        ws.title = "Amazon Reviews"

        # Define headers
        headers = ['Product', 'Reviewer Name', 'Rating', 'Date', 'Review Text', 'Helpful Votes']
        for col_num, header in enumerate(headers, 1):
            ws.cell(row=1, column=col_num, value=header)

        # Start from row 2
        row_num = 2

        # Add summary information
        ws.cell(row=row_num, column=1, value='SUMMARY')
        ws.cell(row=row_num, column=2, value=f'Search Term: {search_term}')
        ws.cell(row=row_num, column=4, value=datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
        row_num += 1

        # Process each product
        for i, product in enumerate(products, 1):
            # Add product header
            product_title = product["title"][:50] + "..." if len(product["title"]) > 50 else product["title"]
            ws.cell(row=row_num, column=1, value=f'PRODUCT {i}: {product_title}')
            ws.cell(row=row_num, column=2, value=product["url"])
            ws.cell(row=row_num, column=3, value=f'Reviews: {product["reviews_count"]}')
            ws.cell(row=row_num, column=4, value='Success: Yes' if product["success"] else 'Success: No')
            row_num += 1

            # Add reviews for this product
            reviews = product.get("reviews", [])
            for review in reviews:
                ws.cell(row=row_num, column=1, value=product_title)
                ws.cell(row=row_num, column=2, value=review.get("reviewer_name", ""))
                ws.cell(row=row_num, column=3, value=review.get("rating", ""))
                ws.cell(row=row_num, column=4, value=review.get("date", ""))
                ws.cell(row=row_num, column=5, value=review.get("text", ""))
                ws.cell(row=row_num, column=6, value=review.get("helpful_votes", ""))
                row_num += 1

        # Save to exports directory
        exports_dir = os.path.join(os.path.dirname(__file__), '..', 'exports')
        os.makedirs(exports_dir, exist_ok=True)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"amazon_reviews_{search_term.replace(' ', '_')}_{timestamp}.xlsx"
        filepath = os.path.join(exports_dir, filename)

        # Save the workbook
        wb.save(filepath)

        return filepath, filename

    # ---------------- Helper: Load/Save authentication config ----------------
    def load_auth_config():
        """Load authentication configuration from file"""
        config_file = os.path.join(os.path.dirname(__file__), 'auth_config.json')
        if os.path.exists(config_file):
            try:
                with open(config_file, 'r') as f:
                    config = json.load(f)
                    # Merge with default config
                    merged_config = AUTO_AUTH_CONFIG.copy()
                    merged_config.update(config)
                    return merged_config
            except Exception as e:
                logger.warning(f"⚠️ Error loading auth config: {str(e)}")
        return AUTO_AUTH_CONFIG.copy()

    def save_auth_config(config):
        """Save authentication configuration to file"""
        config_file = os.path.join(os.path.dirname(__file__), 'auth_config.json')
        try:
            with open(config_file, 'w') as f:
                json.dump(config, f, indent=2)
            logger.info("✅ Auth config saved successfully")
            return True
        except Exception as e:
            logger.error(f"❌ Error saving auth config: {str(e)}")
            return False

    def load_session_cookies():
        """Load saved session cookies from file"""
        auth_config = load_auth_config()
        session_file = os.path.join(os.path.dirname(__file__), auth_config['session_file'])
        if os.path.exists(session_file):
            try:
                with open(session_file, 'r') as f:
                    session_data = json.load(f)
                    # Check if cookies are recent (within 24 hours)
                    timestamp = session_data.get('timestamp')
                    if timestamp:
                        from datetime import datetime, timedelta
                        saved_time = datetime.fromisoformat(timestamp.replace('Z', '+00:00').replace('+00:00', ''))
                        if datetime.now() - saved_time > timedelta(hours=24):
                            logger.warning("⚠️ Session cookies are older than 24 hours, will need fresh login")
                            return []
                    logger.info("✅ Loaded existing session cookies")
                    return session_data.get('cookies', [])
            except Exception as e:
                logger.warning(f"⚠️ Error loading session cookies: {str(e)}")
        return []

    def save_session_cookies(cookies):
        """Save current session cookies to file"""
        auth_config = load_auth_config()
        session_file = os.path.join(os.path.dirname(__file__), auth_config['session_file'])
        try:
            session_data = {
                'cookies': cookies,
                'timestamp': datetime.now().isoformat()
            }
            with open(session_file, 'w') as f:
                json.dump(session_data, f, indent=2)
            logger.info("✅ Session cookies saved successfully")
            return True
        except Exception as e:
            logger.error(f"❌ Error saving session cookies: {str(e)}")
            return False

    # ---------------- Helper: Amazon Auto Login ----------------
    async def auto_login_amazon(page):
        """Automatically log in to Amazon using stored credentials"""
        auth_config = load_auth_config()
        
        if not auth_config['enabled']:
            logger.info("🔓 Auto auth is disabled, continuing as anonymous user")
            return False
            
        if not auth_config['credentials']['email'] or not auth_config['credentials']['password']:
            logger.warning("⚠️ Auto auth enabled but credentials not set")
            return False

        try:
            logger.info("🔐 Starting auto login to Amazon...")
            
            # Quick check if already logged in by looking for account info
            try:
                account_indicator = page.locator('#nav-link-accountList')
                if await account_indicator.count() > 0:
                    account_text = await account_indicator.text_content()
                    if account_text and 'Hello' in account_text and 'sign in' not in account_text.lower():
                        logger.info("✅ Already logged in to Amazon")
                        return True
            except:
                pass

            # Navigate to Amazon homepage first
            logger.info("🌐 Navigating to Amazon homepage...")
            await page.goto("https://www.amazon.com", timeout=20000)
            await page.wait_for_load_state("domcontentloaded")
            await human_delay(2000, 3000)
            
            # Add some human-like behavior
            await human_scroll(page, max_scrolls=1)
            await human_delay(1000, 2000)
            
            # Try multiple approaches to find and click sign-in
            signin_clicked = False
            signin_selectors = [
                '#nav-link-accountList',
                'a[href*="signin"]',
                'a[href*="ap/signin"]',
                '.nav-line-1',
                '#nav-link-accountList .nav-line-1',
                '[data-nav-id="nav_ya_signin"]',
                '.nav-line-2',
                '#nav-link-accountList .nav-line-2',
                'a[href*="homepage.html"]',  # Your Account link
                'text=Your Account',
                'text=Account & Lists',
                'text=Hello, Sign in'
            ]
            
            logger.info("🔍 Looking for sign-in elements...")
            for selector in signin_selectors:
                try:
                    signin_element = page.locator(selector)
                    if await signin_element.count() > 0:
                        # Get element text for debugging
                        element_text = await signin_element.text_content()
                        element_href = await signin_element.get_attribute('href')
                        logger.info(f"✅ Found sign-in element '{selector}': text='{element_text}', href='{element_href}'")
                        
                        # Try to click it
                        await signin_element.click()
                        logger.info(f"✅ Clicked sign-in element: {selector}")
                        await page.wait_for_load_state("domcontentloaded")
                        await human_delay(2000, 3000)
                        signin_clicked = True
                        break
                except Exception as e:
                    logger.warning(f"⚠️ Could not click sign-in with selector {selector}: {str(e)}")
                    continue
            
            if not signin_clicked:
                # Try alternative approach - look for "Account & Lists" or similar
                logger.info("🔍 Trying alternative sign-in approach...")
                alt_selectors = [
                    'text=Account & Lists',
                    'text=Hello, Sign in',
                    'text=Sign in',
                    '[data-nav-id="nav_ya_signin"]',
                    '.nav-line-1:has-text("Hello")',
                    '.nav-line-2:has-text("Sign in")'
                ]
                
                for selector in alt_selectors:
                    try:
                        element = page.locator(selector)
                        if await element.count() > 0:
                            await element.click()
                            logger.info(f"✅ Clicked alternative sign-in element: {selector}")
                            await page.wait_for_load_state("domcontentloaded")
                            await human_delay(2000, 3000)
                            signin_clicked = True
                            break
                    except Exception as e:
                        logger.warning(f"⚠️ Could not click alternative selector {selector}: {str(e)}")
                        continue
            
            # Check for error pages after navigation
            page_content = await page.content()
            if "Looking for Something?" in page_content or "We're sorry" in page_content:
                logger.error("❌ Amazon shows error page - may be blocking automated access")
                # Try a different approach - go to a specific product page first
                logger.info("🔄 Trying workaround - navigating to a product page first...")
                await page.goto("https://www.amazon.com/dp/B08N5WRWNW", timeout=20000)  # Popular product
                await page.wait_for_load_state("domcontentloaded")
                await human_delay(2000, 3000)
                
                # Now try to find sign-in from product page
                for selector in signin_selectors:
                    try:
                        signin_element = page.locator(selector)
                        if await signin_element.count() > 0:
                            await signin_element.click()
                            logger.info(f"✅ Clicked sign-in from product page: {selector}")
                            await page.wait_for_load_state("domcontentloaded")
                            await human_delay(2000, 3000)
                            signin_clicked = True
                            break
                    except Exception as e:
                        continue
                
                # Check again for errors
                page_content = await page.content()
                if "Looking for Something?" in page_content:
                    logger.error("❌ Still getting error page - Amazon may be blocking automated access")
                    return False
            
            # Check if we're actually on a sign-in page
            if not signin_clicked:
                logger.error("❌ Could not find or click any sign-in elements")
                return False
            
            # Wait for sign-in page to load properly - use domcontentloaded instead of networkidle
            await page.wait_for_load_state("domcontentloaded")
            await human_delay(2000, 3000)
            
            # Final check for sign-in page
            page_content = await page.content()
            if "sign in" not in page_content.lower() and "email" not in page_content.lower():
                logger.error("❌ Not on Amazon sign-in page after navigation")
                return False
            
            # Wait for email input field with explicit wait
            logger.info("🔍 Waiting for email input field...")
            email_selectors = [
                'input[name="email"]',
                'input[id="ap_email"]', 
                'input[type="email"]',
                'input[placeholder*="email" i]',
                'input[placeholder*="Email" i]'
            ]
            
            email_filled = False
            for email_selector in email_selectors:
                try:
                    # Wait for element to be visible and fillable
                    await page.wait_for_selector(email_selector, timeout=10000)
                    await page.fill(email_selector, auth_config['credentials']['email'])
                    logger.info(f"✅ Filled email using selector: {email_selector}")
                    email_filled = True
                    break
                except Exception as e:
                    logger.warning(f"⚠️ Could not fill email with selector {email_selector}: {str(e)}")
                    continue
            
            if not email_filled:
                logger.error("❌ Could not find email input field")
                return False
            
            await human_delay(500, 800)
            
            # Click continue button if present
            try:
                continue_selectors = [
                    'input[id="continue"]',
                    'input[type="submit"][value*="Continue"]',
                    'input[type="submit"][value*="continue"]',
                    'button[type="submit"]:has-text("Continue")',
                    'input[aria-labelledby="continue-announce"]'
                ]
                
                for continue_selector in continue_selectors:
                    try:
                        continue_btn = page.locator(continue_selector)
                        if await continue_btn.count() > 0:
                            await continue_btn.click()
                            logger.info(f"✅ Clicked continue button: {continue_selector}")
                            await page.wait_for_load_state("domcontentloaded")
                            await human_delay(1000, 1500)
                            break
                    except:
                        continue
            except Exception as e:
                logger.warning(f"⚠️ Could not click continue button: {str(e)}")

            # Wait for password field and fill it
            logger.info("🔍 Waiting for password input field...")
            password_selectors = [
                'input[name="password"]',
                'input[id="ap_password"]',
                'input[type="password"]',
                'input[placeholder*="password" i]',
                'input[placeholder*="Password" i]'
            ]
            
            password_filled = False
            for password_selector in password_selectors:
                try:
                    # Wait for element to be visible and fillable
                    await page.wait_for_selector(password_selector, timeout=10000)
                    await page.fill(password_selector, auth_config['credentials']['password'])
                    logger.info(f"✅ Filled password using selector: {password_selector}")
                    password_filled = True
                    break
                except Exception as e:
                    logger.warning(f"⚠️ Could not fill password with selector {password_selector}: {str(e)}")
                    continue
            
            if not password_filled:
                logger.error("❌ Could not find password input field")
                return False

            await human_delay(500, 800)

            # Click sign-in button
            logger.info("🔍 Looking for sign-in button...")
            signin_btn_selectors = [
                'input[id="signInSubmit"]',
                'input[type="submit"]',
                'button[type="submit"]',
                'input[value*="Sign in"]',
                'input[value*="sign in"]',
                'button:has-text("Sign in")',
                'button:has-text("sign in")',
                'input[aria-labelledby="signInSubmit-announce"]'
            ]
            
            signin_clicked = False
            for signin_selector in signin_btn_selectors:
                try:
                    signin_btn = page.locator(signin_selector)
                    if await signin_btn.count() > 0:
                        await signin_btn.click()
                        logger.info(f"✅ Clicked sign-in button: {signin_selector}")
                        signin_clicked = True
                        break
                except Exception as e:
                    logger.warning(f"⚠️ Could not click sign-in button with selector {signin_selector}: {str(e)}")
                    continue
            
            if not signin_clicked:
                logger.error("❌ Could not find or click sign-in button")
                return False
            
            # Wait for login to complete with better detection
            logger.info("⏱️ Monitoring for login completion...")
            login_detected = False
            
            # Check for various success indicators
            for i in range(30):  # 30 * 500ms = 15 seconds max
                try:
                    # Check for immediate login success indicators
                    success_indicators = [
                        '#nav-link-accountList:has-text("Hello")',
                        '#nav-link-accountList:not(:has-text("Sign in"))',
                        '[data-nav-id="nav_ya_signin"]:not(:has-text("Sign in"))',
                        '.nav-line-1:has-text("Hello")',
                        '.nav-line-2:not(:has-text("Sign in"))'
                    ]
                    
                    for indicator in success_indicators:
                        try:
                            element = page.locator(indicator)
                            if await element.count() > 0:
                                account_text = await element.text_content()
                                if account_text and 'Hello' in account_text and 'sign in' not in account_text.lower():
                                    logger.info(f"🚀 Login detected! (after {(i+1)*0.5:.1f}s) - {account_text.strip()}")
                                    login_detected = True
                                    break
                        except:
                            continue
                    
                    if login_detected:
                        break
                    
                    # Check for error messages and special cases
                    error_selectors = [
                        '.a-alert-error',
                        '[data-testid="auth-error"]',
                        '.a-box-information .a-alert-heading',
                        '.a-alert-content',
                        '[id*="error"]',
                        '.error-message',
                        '.a-alert-heading',
                        '.a-alert-content'
                    ]
                    
                    for error_selector in error_selectors:
                        try:
                            error_element = page.locator(error_selector)
                            if await error_element.count() > 0:
                                error_text = await error_element.text_content()
                                if error_text and error_text.strip():
                                    logger.warning(f"⚠️ Login error detected: {error_text.strip()}")
                                    # Check for specific error types
                                    error_lower = error_text.lower()
                                    if any(keyword in error_lower for keyword in ['captcha', 'verify', 'robot', 'automated']):
                                        logger.error("❌ CAPTCHA/Verification detected - manual intervention required")
                                        return False
                                    elif any(keyword in error_lower for keyword in ['2fa', 'two-factor', 'verification code', 'sms', 'phone']):
                                        logger.error("❌ 2FA detected - manual intervention required")
                                        return False
                                    elif any(keyword in error_lower for keyword in ['incorrect', 'wrong', 'invalid', 'failed']):
                                        logger.error("❌ Invalid credentials detected")
                                        return False
                                    break
                        except:
                            continue
                    
                    # Check for 2FA or verification pages
                    try:
                        # Look for 2FA indicators
                        twofa_indicators = [
                            'text=Enter the code',
                            'text=verification code',
                            'text=Two-factor',
                            'text=2FA',
                            'input[placeholder*="code" i]',
                            'input[placeholder*="verification" i]'
                        ]
                        
                        for indicator in twofa_indicators:
                            element = page.locator(indicator)
                            if await element.count() > 0:
                                logger.error("❌ 2FA/Verification page detected - manual intervention required")
                                return False
                    except:
                        pass
                    
                    # Check if we're still on login page (might indicate failure)
                    current_url = page.url
                    if 'signin' not in current_url.lower() and 'ap/signin' not in current_url.lower():
                        # We might have been redirected, check if we're logged in
                        try:
                            account_indicator = page.locator('#nav-link-accountList')
                            if await account_indicator.count() > 0:
                                account_text = await account_indicator.text_content()
                                if account_text and 'Hello' in account_text:
                                    logger.info(f"🚀 Login detected via URL change! (after {(i+1)*0.5:.1f}s)")
                                    login_detected = True
                                    break
                        except:
                            pass
                
                except Exception as e:
                    logger.warning(f"⚠️ Error during login detection: {str(e)}")
                
                await human_delay(500, 500)  # Wait 500ms before next check
            
            # Final verification
            if login_detected:
                logger.info("✅ Successfully logged in to Amazon")
                # Save session cookies if persistent session is enabled
                if auth_config['persistent_session']:
                    cookies = await page.context.cookies()
                    save_session_cookies(cookies)
                return True
            
            # One final comprehensive check
            try:
                # Check multiple indicators one more time
                account_indicator = page.locator('#nav-link-accountList')
                if await account_indicator.count() > 0:
                    account_text = await account_indicator.text_content()
                    if account_text and ('Hello' in account_text or 'Account' in account_text) and 'sign in' not in account_text.lower():
                        logger.info("✅ Login successful (final verification)")
                        if auth_config['persistent_session']:
                            cookies = await page.context.cookies()
                            save_session_cookies(cookies)
                        return True
            except:
                pass

            logger.warning("⚠️ Login may have failed - could not verify account status")
            logger.info("💡 Tip: You can disable authentication in the settings to continue without login")
            return False

        except Exception as e:
            logger.error(f"❌ Auto login failed: {str(e)}")
            return False

    # ---------------- Helper: Human-like delay ----------------
    async def human_delay(min_ms=1000, max_ms=3000):
        await asyncio.sleep(random.uniform(min_ms, max_ms) / 1000)
    
    # ---------------- Helper: Safe page navigation with timeout handling ----------------
    async def safe_navigate(page, url, timeout=60000, max_retries=2):
        """Navigate to URL with retry logic and better error handling"""
        for attempt in range(max_retries + 1):
            try:
                logger.info(f"🌐 Navigating to {url} (attempt {attempt + 1}/{max_retries + 1})")
                await page.goto(url, timeout=timeout)
                await page.wait_for_load_state("domcontentloaded")
                logger.info(f"✅ Successfully loaded {url}")
                return True
            except Exception as e:
                logger.warning(f"⚠️ Navigation attempt {attempt + 1} failed: {str(e)}")
                if attempt < max_retries:
                    await human_delay(2000, 4000)  # Wait before retry
                    continue
                else:
                    logger.error(f"❌ All navigation attempts failed for {url}")
                    return False
        return False

    # ---------------- Helper: Human-like scrolling ----------------
    async def human_scroll(page, max_scrolls=3):
        """Perform human-like scrolling to load dynamic content and appear more natural"""
        try:
            # Get page height
            page_height = await page.evaluate("document.body.scrollHeight")

            for i in range(random.randint(1, max_scrolls)):
                # Random scroll amount (300-800px)
                scroll_amount = random.randint(300, 800)
                current_scroll = await page.evaluate("window.pageYOffset")

                # Sometimes use smooth scrolling, sometimes instant
                if random.choice([True, False]):
                    # Smooth scroll
                    await page.evaluate(f"""
                        window.scrollTo({{
                            top: {current_scroll + scroll_amount},
                            behavior: 'smooth'
                        }});
                    """)
                    await human_delay(800, 1500)  # Wait for smooth scroll
                else:
                    # Instant scroll
                    await page.evaluate(f"window.scrollTo(0, {current_scroll + scroll_amount})")
                    await human_delay(300, 800)  # Shorter wait for instant scroll

                # Random pause to simulate reading
                if random.random() < 0.3:  # 30% chance
                    await human_delay(1000, 3000)

            # Sometimes scroll back up a bit (20% chance)
            if random.random() < 0.2:
                back_scroll = random.randint(100, 300)
                current_scroll = await page.evaluate("window.pageYOffset")
                await page.evaluate(f"window.scrollTo(0, {max(0, current_scroll - back_scroll)})")
                await human_delay(500, 1000)

            logger.info(f"🔄 Performed {max_scrolls} scroll actions to load dynamic content")

        except Exception as e:
            logger.warning(f"⚠️ Scrolling failed: {str(e)}")

    # ---------------- Helper: Extract product reviews from multiple pages ----------------
    async def extract_product_reviews(product_url, max_reviews=50, max_pages=3):
        """Extract customer reviews from multiple pages - first tries product page reviews, then navigates to reviews page and extracts from up to 3 pages"""
        browser = None
        playwright = None
        page = None
        try:
            browser, playwright, page = await get_browser(headless=False)
            logger.info(f"📝 STARTING MULTI-PAGE REVIEW EXTRACTION")
            logger.info(f"   🎯 Target: {product_url}")
            logger.info(f"   📊 Max pages to scrape: {max_pages}")
            logger.info(f"   📈 Max reviews per page: {max_reviews}")
            # Use safe navigation with retry logic
            if not await safe_navigate(page, product_url, timeout=60000):
                logger.error(f"❌ Failed to navigate to product URL: {product_url}")
                return {
                    "url": product_url,
                    "reviews": [],
                    "success": False,
                    "error": "Failed to load product page",
                    "pages_scraped": 0
                }
            await human_delay(2000, 4000)

            # Add human-like scrolling to load dynamic content
            await human_scroll(page, max_scrolls=2)

            # Helper function to extract reviews from soup
            def extract_reviews_from_soup(soup, max_reviews=None):
                reviews = []
                review_containers = []

                # Try multiple selectors for review containers
                review_selectors = [
                    'li[data-hook="review"]',
                    '.review',
                    '[data-hook="review"]',
                    '.a-section.review'
                ]

                for selector in review_selectors:
                    containers = soup.select(selector)
                    if containers:
                        logger.info(f"✅ Found {len(containers)} review containers with selector: {selector}")
                        review_containers = containers
                        break

                if not review_containers:
                    logger.warning("⚠️ No review containers found")
                    return reviews

                logger.info(f"🔍 Found {len(review_containers)} review containers")

                # Limit reviews if max_reviews is specified
                containers_to_process = review_containers
                if max_reviews is not None:
                    containers_to_process = review_containers[:max_reviews]

                for i, review in enumerate(containers_to_process):
                    try:
                        # Extract review text
                        review_text = ""
                        text_selectors = [
                            '[data-hook="review-body"]',
                            '.review-text-content',
                            '.a-expander-content',
                            '[data-hook="review-collapsed"]'
                        ]
                        for selector in text_selectors:
                            text_el = review.select_one(selector)
                            if text_el:
                                review_text = text_el.get_text(strip=True)
                                break

                        # Extract rating
                        rating = ""
                        rating_selectors = [
                            '[data-hook="review-star-rating"] .a-icon-alt',
                            '.a-icon-star .a-icon-alt',
                            '[data-hook="cmps-review-star-rating"]'
                        ]
                        for selector in rating_selectors:
                            rating_el = review.select_one(selector)
                            if rating_el:
                                rating_text = rating_el.get_text(strip=True)
                                import re
                                match = re.search(r"(\d+\.?\d*)", rating_text)
                                if match:
                                    rating = match.group(1)
                                    break

                        # Extract reviewer name
                        reviewer_name = ""
                        name_selectors = [
                            '[data-hook="review-author"]',
                            '.a-profile-name',
                            '[data-hook="cmps-reviewer-name"]'
                        ]
                        for selector in name_selectors:
                            name_el = review.select_one(selector)
                            if name_el:
                                reviewer_name = name_el.get_text(strip=True)
                                break

                        # Extract review date
                        review_date = ""
                        date_selectors = [
                            '[data-hook="review-date"]',
                            '.review-date',
                            '[data-hook="cmps-review-date"]'
                        ]
                        for selector in date_selectors:
                            date_el = review.select_one(selector)
                            if date_el:
                                review_date = date_el.get_text(strip=True)
                                break

                        # Extract helpful votes
                        helpful_votes = ""
                        helpful_selectors = [
                            '[data-hook="helpful-vote-statement"]',
                            '.helpful-votes',
                            '[data-hook="cmps-helpful-vote-statement"]'
                        ]
                        for selector in helpful_selectors:
                            helpful_el = review.select_one(selector)
                            if helpful_el:
                                helpful_votes = helpful_el.get_text(strip=True)
                                break

                        if review_text:  # Only add reviews that have content
                            reviews.append({
                                "reviewer_name": reviewer_name,
                                "rating": rating,
                                "date": review_date,
                                "text": review_text,
                                "helpful_votes": helpful_votes
                            })

                    except Exception as e:
                        logger.warning(f"⚠️ Error extracting review {i+1}: {str(e)}")
                        continue

                return reviews

            # First, try to extract reviews directly from the product page
            soup = BeautifulSoup(await page.content(), "html.parser")
            # Extract all available reviews from product page
            product_page_reviews = extract_reviews_from_soup(soup)

            # If max_pages is 1, we only want product page reviews
            if max_pages == 1 and product_page_reviews:
                result = {
                    "url": product_url,
                    "total_reviews_found": len(product_page_reviews),
                    "reviews": product_page_reviews,
                    "success": True,
                    "source": "product_page",
                    "pages_scraped": 1
                }
                logger.info(f"✅ Extracted {len(product_page_reviews)} reviews from product page (max_pages=1)")
                return result
            
            # For multi-page extraction (max_pages > 1), always navigate to dedicated reviews page
            logger.info(f"🔄 Multi-page extraction requested ({max_pages} pages), navigating to dedicated reviews page...")
            if product_page_reviews:
                logger.info(f"📋 Found {len(product_page_reviews)} reviews on product page, but proceeding to reviews page for multi-page extraction")

            # Navigate to dedicated reviews page for multi-page extraction
            reviews_loaded = False

            # Method 1: Skip clicking review link due to strict mode violations
            # (Multiple elements with same ID cause Playwright strict mode errors)
            logger.info("⏭️ Skipping review link click to avoid strict mode violations")

            # Method 2: Always try direct navigation for multi-page extraction (more reliable)
            if not reviews_loaded or max_pages > 1:
                if max_pages > 1:
                    logger.info(f"🎯 Forcing navigation to reviews page for multi-page extraction (max_pages={max_pages})")
                
                # Check if we're already on a reviews page
                if "/product-reviews/" in product_url:
                    logger.info("✅ Already on a reviews page, no need to navigate")
                    reviews_loaded = True
                else:
                    try:
                        # Extract product ID from URL
                        import re
                        product_id_match = re.search(r'/dp/([A-Z0-9]+)', product_url)
                        if product_id_match:
                            product_id = product_id_match.group(1)
                            reviews_url = f"https://www.amazon.com/product-reviews/{product_id}"
                            logger.info(f"🔄 Navigating directly to reviews page: {reviews_url}")

                            await page.goto(reviews_url, timeout=60000)
                            await page.wait_for_load_state("domcontentloaded")
                            await human_delay(3000, 5000)  # Increased wait time

                            # Wait for DOM to be ready (JavaScript loading reviews)
                            await page.wait_for_load_state("domcontentloaded")
                            await human_delay(2000, 3000)

                            # Try to wait for review elements to appear with longer timeout
                            try:
                                await page.wait_for_selector('[data-hook="review"], .review, .a-section.review', timeout=SELECTOR_TIMEOUT)
                                logger.info("✅ Review elements found on page")
                            except Exception as e:
                                logger.warning(f"⚠️ Review elements not found within timeout: {str(e)}")
                                # Continue anyway - some pages might have different selectors

                            await human_scroll(page, max_scrolls=3)  # More scrolling to trigger loading
                            await human_delay(1000, 2000)

                            reviews_loaded = True
                            logger.info("✅ Navigated to reviews page directly")
                        else:
                            logger.warning("⚠️ Could not extract product ID for direct navigation")

                    except Exception as e:
                        logger.warning(f"⚠️ Direct navigation to reviews failed: {str(e)}")

            # Now extract reviews from multiple pages
            all_reviews = []
            pages_scraped = 0

            for page_num in range(1, max_pages + 1):
                try:
                    logger.info(f"📖 SCRAPING PAGE {page_num}/{max_pages} | Current URL: {page.url}")
                    
                    # Wait for page to load completely - use domcontentloaded instead of networkidle
                    await page.wait_for_load_state("domcontentloaded")
                    await human_delay(2000, 3000)
                    
                    # Scroll to load dynamic content
                    logger.info(f"🔄 Scrolling page {page_num} to load dynamic content...")
                    await human_scroll(page, max_scrolls=2)
                    await human_delay(1000, 2000)

                    soup = BeautifulSoup(await page.content(), "html.parser")

                    # Extract reviews from current page
                    logger.info(f"🔍 Extracting reviews from page {page_num}...")
                    page_reviews = extract_reviews_from_soup(soup)
                    
                    if page_reviews:
                        all_reviews.extend(page_reviews)
                        pages_scraped += 1
                        logger.info(f"✅ PAGE {page_num} SUCCESS: Extracted {len(page_reviews)} reviews | Total so far: {len(all_reviews)}")
                    else:
                        logger.warning(f"⚠️ PAGE {page_num} NO REVIEWS: No reviews found on this page")

                    # If this is not the last page, try to navigate to next page
                    if page_num < max_pages:
                        logger.info(f"🚀 NAVIGATING TO PAGE {page_num + 1}...")
                        # Look for next page button
                        next_page_found = False
                        next_selectors = [
                            'li.a-last a',
                            'a[aria-label="Next page"]',
                            'li.a-last:not(.a-disabled) a',
                            '.a-pagination .a-last a',
                            'text=Next',
                            '[data-hook="pagination-bar"] .a-last a'
                        ]

                        for selector in next_selectors:
                            try:
                                next_button = page.locator(selector)
                                if await next_button.count() > 0:
                                    # Check if button is enabled (not disabled)
                                    button_class = await next_button.first.get_attribute('class') or ""
                                    parent_class = await next_button.first.locator('..').get_attribute('class') or ""
                                    
                                    if 'a-disabled' not in button_class and 'a-disabled' not in parent_class:
                                        logger.info(f"🔗 PAGE {page_num}: Found next page button with selector: {selector}")
                                        await next_button.first.click()
                                        logger.info(f"⏳ PAGE {page_num}: Clicked next button, waiting for page {page_num + 1} to load...")
                                        await human_delay(3000, 5000)  # Wait for page to load
                                        logger.info(f"✅ PAGE {page_num}: Successfully navigated to page {page_num + 1}")
                                        next_page_found = True
                                        break
                            except Exception as e:
                                logger.warning(f"⚠️ PAGE {page_num}: Error trying next page selector {selector}: {str(e)}")
                                continue

                        if not next_page_found:
                            logger.info(f"🔚 PAGE {page_num}: No more pages available after page {page_num} (reached end or no next button)")
                            break

                except Exception as e:
                    logger.error(f"❌ Error scraping page {page_num}: {str(e)}")
                    break

            result = {
                "url": product_url,
                "total_reviews_found": len(all_reviews),
                "reviews": all_reviews,
                "success": True,
                "source": "reviews_page",
                "pages_scraped": pages_scraped
            }

            logger.info(f"🎯 FINAL SUMMARY: Extracted {len(all_reviews)} total reviews from {pages_scraped}/{max_pages} pages")
            return result

        except Exception as e:
            logger.error(f"❌ Error getting product reviews for {product_url}: {str(e)}")
            return {
                "url": product_url,
                "reviews": [],
                "success": False,
                "error": str(e),
                "pages_scraped": 0
            }
        finally:
            # Safely close browser and playwright
            try:
                if browser:
                    await browser.close()
                if playwright:
                    await playwright.stop()
            except Exception as e:
                logger.warning(f"⚠️ Error closing browser/playwright: {str(e)}")

    # ---------------- Helper: Extract product details from page ----------------
    async def extract_product_details(product_url):
        """Extract product details from a single product page"""
        browser, playwright, page = await get_browser(headless=False)
        try:
            logger.info(f"📦 Getting details for: {product_url}")
            await page.goto(product_url, timeout=60000)
            await page.wait_for_load_state("domcontentloaded")
            await human_delay(2000, 4000)

            # Add human-like scrolling to load dynamic content
            await human_scroll(page, max_scrolls=2)

            soup = BeautifulSoup(await page.content(), "html.parser")

            # Extract title
            title = ""
            title_selectors = [
                "#productTitle",
                "#title",
                ".a-size-large.product-title-word-break",
                "h1.a-size-large"
            ]
            for selector in title_selectors:
                title_el = soup.select_one(selector)
                if title_el:
                    title = title_el.get_text(strip=True)
                    break

            # Extract price
            price = ""
            price_selectors = [
                ".a-price .a-offscreen",
                "#priceblock_ourprice",
                "#priceblock_dealprice",
                "#priceblock_saleprice",
                ".a-price-whole",
                ".a-color-price",
                "#corePrice_feature_div .a-price .a-offscreen",
                "#corePriceDisplay_desktop_feature_div .a-price-whole",
                "#corePriceDisplay_desktop_feature_div .a-price-fraction",
                ".a-price .a-offscreen:first-child",
                "#snsPrice .a-price .a-offscreen",
                ".apexPriceToPay .a-offscreen"
            ]
            for selector in price_selectors:
                price_el = soup.select_one(selector)
                if price_el:
                    price_text = price_el.get_text(strip=True)
                    # Clean up price text
                    if price_text and ("$" in price_text or "CDN$" in price_text):
                        price = price_text
                        break

            # Extract rating
            rating = ""
            rating_selectors = [
                ".a-icon-star-small .a-icon-alt",
                ".a-icon-star .a-icon-alt",
                "#acrPopover .a-icon-alt",
                ".review-rating",
                "#averageCustomerReviews .a-icon-alt",
                ".a-declarative[data-action='acrStarsLink'] .a-icon-alt",
                "#acrPopover [data-cy='reviews-ratings-date']"
            ]
            for selector in rating_selectors:
                rating_el = soup.select_one(selector)
                if rating_el:
                    rating_text = rating_el.get_text(strip=True)
                    # Extract number from "4.5 out of 5 stars"
                    import re
                    match = re.search(r"(\d+\.?\d*)", rating_text)
                    if match:
                        rating = match.group(1)
                        break

            # Extract review count
            review_count = ""
            review_selectors = [
                "#acrCustomerReviewText",
                ".a-size-base",
                "a[href*='customerReviews']",
                "#acrCustomerReviewLink",
                "[data-cy='reviews-ratings-count']",
                ".a-link-emphasis"
            ]
            for selector in review_selectors:
                review_el = soup.select_one(selector)
                if review_el:
                    review_text = review_el.get_text(strip=True)
                    # Look for patterns like "1,234 ratings" or "1,234 customer reviews"
                    import re
                    match = re.search(r"([\d,]+)", review_text)
                    if match:
                        review_count = match.group(1)
                        break

            result = {
                "url": product_url,
                "title": title,
                "price": price,
                "rating": rating,
                "review_count": review_count,
                "success": True
            }

            logger.info(f"✅ Extracted details: {title[:50]}... | Price: {price} | Rating: {rating} | Reviews: {review_count}")
            return result

        except Exception as e:
            logger.error(f"❌ Error getting product details for {product_url}: {str(e)}")
            return {
                "url": product_url,
                "title": "",
                "price": "",
                "rating": "",
                "review_count": "",
                "success": False,
                "error": str(e)
            }
        finally:
            await browser.close()
            await playwright.stop()

    # ---------------- Helper: Start browser with rotation and auto auth ----------------
    async def get_browser(headless=False, auto_login=True):  # Changed to non-headless for testing
        ua = random.choice(USER_AGENTS)
        viewport = {
            "width": random.randint(1280, 1920),
            "height": random.randint(720, 1080),
        }

        logger.info(f"🌐 Using UA: {ua[:60]}..., viewport={viewport}")

        playwright = await async_playwright().start()
        browser = await playwright.chromium.launch(
            headless=headless,
            args=[
                "--disable-blink-features=AutomationControlled",
                "--disable-dev-shm-usage",
                "--no-sandbox",
                "--disable-setuid-sandbox",
                "--disable-web-security",
                "--disable-extensions",
                "--disable-plugins",
                "--disable-background-timer-throttling",
                "--disable-backgrounding-occluded-windows",
                "--disable-renderer-backgrounding",
                "--disable-features=VizDisplayCompositor",
                "--disable-ipc-flooding-protection",
                "--disable-hang-monitor",
                "--disable-prompt-on-repost",
                "--disable-sync",
                "--disable-translate",
                "--disable-logging",
                "--disable-permissions-api",
                "--disable-notifications",
                "--disable-geolocation",
                "--disable-speech-api",
                "--disable-file-system",
                "--disable-presentation-api",
                "--disable-remote-fonts",
                "--disable-client-side-phishing-detection",
                "--disable-component-extensions-with-background-pages",
                "--disable-default-apps",
                "--disable-domain-reliability",
                "--disable-features=TranslateUI",
                "--hide-scrollbars",
                "--mute-audio",
                "--no-first-run",
                "--no-default-browser-check",
                "--no-pings",
                "--no-zygote",
                "--use-mock-keychain",
                "--user-agent=" + ua,
            ]
        )
        
        context = await browser.new_context(
            user_agent=ua,
            locale="en-US",
            timezone_id="America/New_York",
            viewport=viewport,
            ignore_https_errors=True,
            extra_http_headers={
                "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8",
                "Accept-Language": "en-US,en;q=0.5",
                "Accept-Encoding": "gzip, deflate, br",
                "DNT": "1",
                "Connection": "keep-alive",
                "Upgrade-Insecure-Requests": "1",
                "Sec-Fetch-Dest": "document",
                "Sec-Fetch-Mode": "navigate",
                "Sec-Fetch-Site": "none",
                "Cache-Control": "max-age=0"
            }
        )
        
        # Set default timeouts for all operations
        context.set_default_timeout(DEFAULT_TIMEOUT)  # 60 seconds for all operations
        context.set_default_navigation_timeout(NAVIGATION_TIMEOUT)  # 60 seconds for navigation

        # Load saved session cookies if available
        auth_config = load_auth_config()
        cookies_loaded = False
        if auth_config['enabled'] and auth_config['persistent_session']:
            saved_cookies = load_session_cookies()
            if saved_cookies:
                try:
                    await context.add_cookies(saved_cookies)
                    logger.info("✅ Restored session cookies")
                    cookies_loaded = True
                except Exception as e:
                    logger.warning(f"⚠️ Error restoring cookies: {str(e)}")

        # Add more stealth measures
        page = await context.new_page()
        await page.evaluate("""
            // Remove webdriver property
            Object.defineProperty(navigator, 'webdriver', {
                get: () => undefined,
            });
            
            // Override the plugins property to use a custom getter
            Object.defineProperty(navigator, 'plugins', {
                get: () => [1, 2, 3, 4, 5],
            });
            
            // Override the languages property to use a custom getter
            Object.defineProperty(navigator, 'languages', {
                get: () => ['en-US', 'en'],
            });
            
            // Override the permissions property to use a custom getter
            const originalQuery = window.navigator.permissions.query;
            window.navigator.permissions.query = (parameters) => (
                parameters.name === 'notifications' ?
                    Promise.resolve({ state: Notification.permission }) :
                    originalQuery(parameters)
            );
            
            // Mock chrome object
            window.chrome = {
                runtime: {},
                loadTimes: function() {},
                csi: function() {},
                app: {}
            };
            
            // Mock webkit object
            window.webkit = {
                messageHandlers: {}
            };
        """)

        # Perform auto login if enabled and requested
        if auto_login and auth_config['enabled']:
            # If we have cookies, do an ultra-quick validation first
            if cookies_loaded:
                logger.info("⚡ Ultra-fast session validation with restored cookies...")
                await page.goto("https://www.amazon.com", timeout=10000)  # Even faster timeout
                # Skip wait_for_load_state for maximum speed
                await human_delay(200, 300)  # Ultra-minimal delay
                
                # Quick check if already logged in
                try:
                    # Use immediate check without waiting
                    account_indicator = page.locator('#nav-link-accountList')
                    if await account_indicator.count() > 0:
                        account_text = await account_indicator.text_content()
                        if account_text and 'Hello' in account_text and 'sign in' not in account_text.lower():
                            logger.info("🚀 Lightning authentication: Already logged in with saved session!")
                            return browser, playwright, page
                except:
                    pass
                
                logger.info("⚠️ Saved session invalid, proceeding with fresh login...")
            
            # Navigate to Amazon first to establish domain context with faster timeout
            await page.goto("https://www.amazon.com", timeout=15000)  # Reduced from 30s to 15s
            await page.wait_for_load_state("domcontentloaded")
            await human_delay(500, 800)  # Reduced from 1-2s to 0.5-0.8s
            
            # Attempt auto login
            login_success = await auto_login_amazon(page)
            if login_success:
                logger.info("🎉 Auto authentication completed successfully")
            else:
                logger.info("🔓 Continuing without authentication")

        return browser, playwright, page



    # ---------------- Route: Get product details ----------------
    @app.route("/product-details")
    def product_details():
        product_url = request.args.get("url", "").strip()
        if not product_url:
            return jsonify({"error": "Missing product URL", "success": False}), 400

        # Basic validation for Amazon URLs
        if not ("amazon.com" in product_url and ("/dp/" in product_url or "/gp/product/" in product_url or "/product-reviews/" in product_url)):
            return jsonify({"error": "Invalid Amazon product URL", "success": False}), 400

        async def run():
            browser, playwright, page = await get_browser(headless=False)
            try:
                logger.info(f"📦 Getting details for: {product_url}")
                await page.goto(product_url, timeout=60000)
                await page.wait_for_load_state("domcontentloaded")
                await human_delay(2000, 4000)

                # Add human-like scrolling to load dynamic content
                await human_scroll(page, max_scrolls=2)

                soup = BeautifulSoup(await page.content(), "html.parser")

                # Debug: Log some HTML content to identify selectors
                logger.info(f"🔍 Page title: {soup.title.get_text() if soup.title else 'No title'}")
                logger.info(f"🔍 First few price-related elements: {len(soup.select('.a-price'))}")
                logger.info(f"🔍 First few rating elements: {len(soup.select('.a-icon-star'))}")
                logger.info(f"🔍 Sample HTML around price: {str(soup.select_one('.a-price'))[:200] if soup.select_one('.a-price') else 'No .a-price found'}")

                # Extract title
                title = ""
                title_selectors = [
                    "#productTitle",
                    "#title",
                    ".a-size-large.product-title-word-break",
                    "h1.a-size-large"
                ]
                for selector in title_selectors:
                    title_el = soup.select_one(selector)
                    if title_el:
                        title = title_el.get_text(strip=True)
                        break

                # Extract price
                price = ""
                price_selectors = [
                    ".a-price .a-offscreen",
                    "#priceblock_ourprice",
                    "#priceblock_dealprice",
                    "#priceblock_saleprice",
                    ".a-price-whole",
                    ".a-color-price",
                    "#corePrice_feature_div .a-price .a-offscreen",
                    "#corePriceDisplay_desktop_feature_div .a-price-whole",
                    "#corePriceDisplay_desktop_feature_div .a-price-fraction",
                    ".a-price .a-offscreen:first-child",
                    "#snsPrice .a-price .a-offscreen",
                    ".apexPriceToPay .a-offscreen"
                ]
                for selector in price_selectors:
                    price_el = soup.select_one(selector)
                    if price_el:
                        price_text = price_el.get_text(strip=True)
                        # Clean up price text
                        if price_text and ("$" in price_text or "CDN$" in price_text):
                            price = price_text
                            break

                # Extract rating
                rating = ""
                rating_selectors = [
                    ".a-icon-star-small .a-icon-alt",
                    ".a-icon-star .a-icon-alt",
                    "#acrPopover .a-icon-alt",
                    ".review-rating",
                    "#averageCustomerReviews .a-icon-alt",
                    ".a-declarative[data-action='acrStarsLink'] .a-icon-alt",
                    "#acrPopover [data-cy='reviews-ratings-date']"
                ]
                for selector in rating_selectors:
                    rating_el = soup.select_one(selector)
                    if rating_el:
                        rating_text = rating_el.get_text(strip=True)
                        # Extract number from "4.5 out of 5 stars"
                        import re
                        match = re.search(r"(\d+\.?\d*)", rating_text)
                        if match:
                            rating = match.group(1)
                            break

                # Extract review count
                review_count = ""
                review_selectors = [
                    "#acrCustomerReviewText",
                    ".a-size-base",
                    "a[href*='customerReviews']",
                    "#acrCustomerReviewLink",
                    "[data-cy='reviews-ratings-count']",
                    ".a-link-emphasis"
                ]
                for selector in review_selectors:
                    review_el = soup.select_one(selector)
                    if review_el:
                        review_text = review_el.get_text(strip=True)
                        # Look for patterns like "1,234 ratings" or "1,234 customer reviews"
                        import re
                        match = re.search(r"([\d,]+)", review_text)
                        if match:
                            review_count = match.group(1)
                            break

                result = {
                    "url": product_url,
                    "title": title,
                    "price": price,
                    "rating": rating,
                    "review_count": review_count,
                    "success": True
                }

                logger.info(f"✅ Extracted details: {title[:50]}... | Price: {price} | Rating: {rating} | Reviews: {review_count}")
                return result

            except Exception as e:
                logger.error(f"❌ Error getting product details: {str(e)}")
                return {
                    "url": product_url,
                    "error": f"Failed to extract product details: {str(e)}",
                    "success": False
                }
            finally:
                await browser.close()
                await playwright.stop()

        loop = asyncio.new_event_loop()
        asyncio.set_event_loop(loop)
        return loop.run_until_complete(run())

    # ---------------- Route: Search for top 3 product links ----------------
    @app.route("/search")
    def search():
        keyword = (request.args.get("q") or "").strip()
        if not keyword:
            return jsonify({"error": "Missing search keyword"}), 400

        async def run():
            browser, playwright, page = await get_browser(headless=True)
            try:
                # Search Amazon
                search_url = f"https://www.amazon.com/s?k={keyword.replace(' ', '+')}"
                logger.info(f"🔍 Searching for: {keyword}")
                await page.goto(search_url, timeout=60000)
                await page.wait_for_load_state("domcontentloaded")
                await human_delay(2000, 4000)

                # Add scrolling to load more search results
                await human_scroll(page, max_scrolls=1)

                try:
                    await page.wait_for_selector('[data-component-type="s-search-result"]', timeout=15000)
                except:
                    html = await page.content()
                    logger.error("❌ No search results loaded")
                    return {"error": "No search results loaded"}

                soup = BeautifulSoup(await page.content(), "html.parser")
                product_containers = (
                    soup.select('[data-component-type="s-search-result"]')
                    or soup.select('[data-asin][data-index]')
                    or soup.select('.s-result-item')
                )

                logger.info(f"🔍 Found {len(product_containers)} product containers")

                products = []
                for i, container in enumerate(product_containers[:3]):  # Only top 3
                    logger.info(f"Processing container {i+1}: {str(container)[:200]}...")

                    # Try multiple selectors for title and link
                    title_el = (
                        container.select_one('h2 a span') or
                        container.select_one('h2 span') or
                        container.select_one('.a-text-normal') or
                        container.select_one('span.a-text-normal')
                    )

                    link_el = (
                        container.select_one('h2 a') or
                        container.select_one('a.a-link-normal') or
                        container.select_one('a[href*="/dp/"]')
                    )

                    if not (title_el and link_el):
                        logger.warning(f"⚠️ Container {i+1}: Missing title or link")
                        continue

                    product_url = link_el.get("href", "")
                    if product_url.startswith("/"):
                        product_url = f"https://www.amazon.com{product_url}"

                    title_text = title_el.get_text(strip=True)
                    products.append({
                        "title": title_text,
                        "url": product_url
                    })
                    logger.info(f"✅ Found product {i+1}: {title_text[:50]}...")

                return {"keyword": keyword, "products": products}

            finally:
                await browser.close()
                await playwright.stop()

        loop = asyncio.new_event_loop()
        asyncio.set_event_loop(loop)
        return loop.run_until_complete(run())

    # ---------------- Route: Search with detailed product information ----------------
    @app.route("/search-detailed")
    def search_detailed():
        keyword = (request.args.get("q") or "").strip()
        if not keyword:
            return jsonify({"error": "Missing search keyword"}), 400

        async def run():
            browser, playwright, page = await get_browser(headless=True)
            try:
                # Search Amazon
                search_url = f"https://www.amazon.com/s?k={keyword.replace(' ', '+')}"
                logger.info(f"🔍 Detailed search for: {keyword}")
                await page.goto(search_url, timeout=60000)
                await page.wait_for_load_state("domcontentloaded")
                await human_delay(2000, 4000)

                # Add scrolling to load more search results
                await human_scroll(page, max_scrolls=1)

                try:
                    await page.wait_for_selector('[data-component-type="s-search-result"]', timeout=15000)
                except:
                    html = await page.content()
                    logger.error("❌ No search results loaded")
                    return {"error": "No search results loaded"}

                soup = BeautifulSoup(await page.content(), "html.parser")
                product_containers = (
                    soup.select('[data-component-type="s-search-result"]')
                    or soup.select('[data-asin][data-index]')
                    or soup.select('.s-result-item')
                )

                logger.info(f"🔍 Found {len(product_containers)} product containers")

                # Extract basic product info (titles and URLs)
                basic_products = []
                for i, container in enumerate(product_containers[:3]):  # Only top 3
                    logger.info(f"Processing container {i+1}: {str(container)[:200]}...")

                    # Try multiple selectors for title and link
                    title_el = (
                        container.select_one('h2 a span') or
                        container.select_one('h2 span') or
                        container.select_one('.a-text-normal') or
                        container.select_one('span.a-text-normal')
                    )

                    link_el = (
                        container.select_one('h2 a') or
                        container.select_one('a.a-link-normal') or
                        container.select_one('a[href*="/dp/"]')
                    )

                    if not (title_el and link_el):
                        logger.warning(f"⚠️ Container {i+1}: Missing title or link")
                        continue

                    product_url = link_el.get("href", "")
                    if product_url.startswith("/"):
                        product_url = f"https://www.amazon.com{product_url}"

                    title_text = title_el.get_text(strip=True)
                    basic_products.append({
                        "title": title_text,
                        "url": product_url
                    })
                    logger.info(f"✅ Found product {i+1}: {title_text[:50]}...")

                # Close search browser
                await browser.close()
                await playwright.stop()

                # Concurrently fetch detailed information for all products
                logger.info(f"🔄 Fetching detailed info for {len(basic_products)} products concurrently...")
                tasks = [extract_product_details(product["url"]) for product in basic_products]
                detailed_results = await asyncio.gather(*tasks, return_exceptions=True)

                # Process results and handle any exceptions
                products = []
                for i, result in enumerate(detailed_results):
                    if isinstance(result, Exception):
                        logger.error(f"❌ Error fetching details for product {i+1}: {str(result)}")
                        # Return basic info if detailed fetch failed
                        products.append({
                            **basic_products[i],
                            "price": "",
                            "rating": "",
                            "review_count": "",
                            "success": False,
                            "error": str(result)
                        })
                    else:
                        products.append(result)

                logger.info(f"✅ Completed detailed search for '{keyword}' - {len(products)} products with full details")
                return {"keyword": keyword, "products": products}

            except Exception as e:
                logger.error(f"❌ Error in detailed search: {str(e)}")
                return {"error": f"Search failed: {str(e)}"}
            finally:
                # Make sure browser is closed even if there's an error
                try:
                    await browser.close()
                    await playwright.stop()
                except:
                    pass

        loop = asyncio.new_event_loop()
        asyncio.set_event_loop(loop)
        return loop.run_until_complete(run())

    # ---------------- Route: Get product reviews ----------------
    @app.route("/product-reviews")
    def product_reviews():
        product_url = request.args.get("url", "").strip()
        max_reviews = request.args.get("max_reviews", "10")
        max_pages = request.args.get("max_pages", "3")
        
        try:
            max_reviews = int(max_reviews)
            if max_reviews > 50:  # Limit to prevent excessive scraping
                max_reviews = 50
        except ValueError:
            max_reviews = 10
            
        try:
            max_pages = int(max_pages)
            if max_pages > 5:  # Limit to prevent excessive scraping
                max_pages = 5
            if max_pages < 1:
                max_pages = 1
        except ValueError:
            max_pages = 3

        if not product_url:
            return jsonify({"error": "Missing product URL", "success": False}), 400

        # Basic validation for Amazon URLs
        if not ("amazon.com" in product_url and ("/dp/" in product_url or "/gp/product/" in product_url or "/product-reviews/" in product_url)):
            return jsonify({"error": "Invalid Amazon product URL", "success": False}), 400

        async def run():
            result = await extract_product_reviews(product_url, max_reviews, max_pages)
            return result

        loop = asyncio.new_event_loop()
        asyncio.set_event_loop(loop)
        return loop.run_until_complete(run())

    # ---------------- Route: Download CSV file for reviews ----------------
    @app.route("/download-csv/<filename>")
    def download_csv(filename):
        """Download CSV file with review data"""
        try:
            # Find the CSV file in exports directory
            exports_dir = os.path.join(os.path.dirname(__file__), '..', 'exports')
            filepath = os.path.join(exports_dir, filename)

            if not os.path.exists(filepath):
                return jsonify({"error": "File not found", "success": False}), 404

            return send_file(
                filepath,
                as_attachment=True,
                download_name=filename,
                mimetype='text/csv'
            )
        except Exception as e:
            logger.error(f"❌ Error downloading CSV file: {str(e)}")
            return jsonify({"error": "Download failed", "success": False}), 500

    # ---------------- Route: Download Excel file for reviews ----------------
    @app.route("/download-excel/<filename>")
    def download_excel(filename):
        """Download Excel file with review data"""
        try:
            # Find the Excel file in exports directory
            exports_dir = os.path.join(os.path.dirname(__file__), '..', 'exports')
            filepath = os.path.join(exports_dir, filename)

            if not os.path.exists(filepath):
                return jsonify({"error": "File not found", "success": False}), 404

            return send_file(
                filepath,
                as_attachment=True,
                download_name=filename,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
        except Exception as e:
            logger.error(f"❌ Error downloading Excel file: {str(e)}")
            return jsonify({"error": "Download failed", "success": False}), 500

    # ---------------- Route: Search for products and get all reviews from first page ----------------
    @app.route("/search-reviews")
    def search_reviews():
        keyword = (request.args.get("q") or "").strip()
        max_products = request.args.get("max_products", "3")
        min_rating = request.args.get("min_rating", "")
        
        try:
            max_products = int(max_products)
            if max_products > 5:  # Limit to prevent excessive scraping
                max_products = 5
            if max_products < 1:
                max_products = 1
        except ValueError:
            max_products = 3
            
        # Parse min_rating if provided
        min_rating_float = None
        if min_rating:
            try:
                min_rating_float = float(min_rating)
                if min_rating_float < 0 or min_rating_float > 5:
                    return jsonify({"error": "Invalid rating filter. Must be between 0 and 5", "success": False}), 400
            except ValueError:
                return jsonify({"error": "Invalid rating filter format", "success": False}), 400

        if not keyword:
            return jsonify({"error": "Missing search keyword", "success": False}), 400

        async def run():
            browser, playwright, page = await get_browser(headless=True)
            try:
                # Search Amazon
                search_url = f"https://www.amazon.com/s?k={keyword.replace(' ', '+')}"
                logger.info(f"🔍 Searching for: {keyword}")
                await page.goto(search_url, timeout=60000)
                await page.wait_for_load_state("domcontentloaded")
                await human_delay(2000, 4000)

                # Add scrolling to load more search results
                await human_scroll(page, max_scrolls=1)

                try:
                    await page.wait_for_selector('[data-component-type="s-search-result"]', timeout=15000)
                except:
                    html = await page.content()
                    logger.error("❌ No search results loaded")
                    return {"error": "No search results loaded", "search_term": keyword, "success": False}

                soup = BeautifulSoup(await page.content(), "html.parser")
                product_containers = (
                    soup.select('[data-component-type="s-search-result"]')
                    or soup.select('[data-asin][data-index]')
                    or soup.select('.s-result-item')
                )

                logger.info(f"🔍 Found {len(product_containers)} product containers")

                # Extract basic product info (titles and URLs)
                basic_products = []
                for i, container in enumerate(product_containers[:max_products]):
                    logger.info(f"Processing container {i+1}: {str(container)[:200]}...")

                    # Try multiple selectors for title and link
                    title_el = (
                        container.select_one('h2 a span') or
                        container.select_one('h2 span') or
                        container.select_one('.a-text-normal') or
                        container.select_one('span.a-text-normal')
                    )

                    link_el = (
                        container.select_one('h2 a') or
                        container.select_one('a.a-link-normal') or
                        container.select_one('a[href*="/dp/"]')
                    )

                    if not (title_el and link_el):
                        logger.warning(f"⚠️ Container {i+1}: Missing title or link")
                        continue

                    product_url = link_el.get("href", "")
                    if product_url.startswith("/"):
                        product_url = f"https://www.amazon.com{product_url}"

                    title_text = title_el.get_text(strip=True)
                    basic_products.append({
                        "title": title_text,
                        "url": product_url
                    })
                    logger.info(f"✅ Found product {i+1}: {title_text[:50]}...")

                # Close search browser
                await browser.close()
                await playwright.stop()

                if not basic_products:
                    return {
                        "search_term": keyword,
                        "total_products": 0,
                        "total_reviews": 0,
                        "products": [],
                        "success": False,
                        "error": "No products found"
                    }

                # First, get detailed product information to check ratings
                logger.info(f"🔄 Fetching detailed product information for {len(basic_products)} products...")
                detail_tasks = [extract_product_details(product["url"]) for product in basic_products]
                detail_results = await asyncio.gather(*detail_tasks, return_exceptions=True)
                
                # Filter products by rating if min_rating is specified
                filtered_products = []
                if min_rating_float is not None:
                    logger.info(f"🔍 Filtering products by minimum rating: {min_rating_float}")
                    for i, result in enumerate(detail_results):
                        if isinstance(result, Exception):
                            logger.warning(f"⚠️ Error getting details for product {i+1}: {str(result)}")
                            # Include products with errors (no rating info) unless we're being strict
                            if min_rating_float <= 0:  # Only include if no minimum rating requirement
                                filtered_products.append(basic_products[i])
                        else:
                            product_rating = result.get("rating", "")
                            if product_rating:
                                try:
                                    rating_float = float(product_rating)
                                    if rating_float >= min_rating_float:
                                        filtered_products.append(basic_products[i])
                                        logger.info(f"✅ Product '{result.get('title', '')[:50]}...' passed rating filter: {rating_float} >= {min_rating_float}")
                                    else:
                                        logger.info(f"❌ Product '{result.get('title', '')[:50]}...' filtered out: {rating_float} < {min_rating_float}")
                                except ValueError:
                                    logger.warning(f"⚠️ Invalid rating format for product: {product_rating}")
                                    # Include products with invalid rating format unless we're being strict
                                    if min_rating_float <= 0:
                                        filtered_products.append(basic_products[i])
                            else:
                                logger.warning(f"⚠️ No rating found for product: {result.get('title', '')[:50]}...")
                                # Include products with no rating unless we're being strict
                                if min_rating_float <= 0:
                                    filtered_products.append(basic_products[i])
                else:
                    # No rating filter, include all products
                    filtered_products = basic_products
                
                logger.info(f"🎯 Rating filter result: {len(filtered_products)}/{len(basic_products)} products passed the filter")
                
                if not filtered_products:
                    return {
                        "search_term": keyword,
                        "total_products": 0,
                        "total_reviews": 0,
                        "products": [],
                        "success": True,
                        "message": f"No products found with rating {min_rating_float}+ stars"
                    }

                # Concurrently fetch reviews for filtered products
                logger.info(f"🔄 Fetching reviews for {len(filtered_products)} filtered products concurrently...")
                tasks = [extract_product_reviews(product["url"], float('inf'), 3) for product in filtered_products]
                review_results = await asyncio.gather(*tasks, return_exceptions=True)

                # Process results and build optimized response
                products = []
                total_reviews = 0

                for i, result in enumerate(review_results):
                    if isinstance(result, Exception):
                        logger.error(f"❌ Error fetching reviews for product {i+1}: {str(result)}")
                        # Add product with no reviews
                        products.append({
                            "title": filtered_products[i]["title"],
                            "url": filtered_products[i]["url"],
                            "reviews_count": 0,
                            "reviews": [],
                            "success": False,
                            "error": str(result)
                        })
                    else:
                        reviews_count = result.get("total_reviews_found", 0)
                        total_reviews += reviews_count
                        products.append({
                            "title": filtered_products[i]["title"],
                            "url": filtered_products[i]["url"],
                            "reviews_count": reviews_count,
                            "reviews": result.get("reviews", []),
                            "success": result.get("success", False)
                        })

                filter_info = f" with rating filter {min_rating_float}+" if min_rating_float is not None else ""
                logger.info(f"✅ Completed search-reviews for '{keyword}'{filter_info} - {len(products)} products with {total_reviews} total reviews")

                # Generate Excel file
                try:
                    excel_filepath, excel_filename = generate_excel_file(keyword, products)
                    excel_download_url = f"http://localhost:5000/download-excel/{excel_filename}"
                    logger.info(f"📊 Generated Excel file: {excel_filename}")
                except Exception as e:
                    logger.error(f"❌ Error generating Excel file: {str(e)}")
                    excel_download_url = None

                return {
                    "search_term": keyword,
                    "total_products": len(products),
                    "total_reviews": total_reviews,
                    "products": products,
                    "excel_download_url": excel_download_url,
                    "success": True
                }

            except Exception as e:
                logger.error(f"❌ Error in search-reviews: {str(e)}")
                return {
                    "search_term": keyword,
                    "error": f"Search failed: {str(e)}",
                    "success": False
                }
            finally:
                # Make sure browser is closed even if there's an error
                try:
                    await browser.close()
                    await playwright.stop()
                except:
                    pass

        loop = asyncio.new_event_loop()
        asyncio.set_event_loop(loop)
        return loop.run_until_complete(run())

    # ---------------- Route: Get authentication config ----------------
    @app.route("/auth-config", methods=["GET"])
    def get_auth_config():
        """Get current authentication configuration (without password)"""
        try:
            config = load_auth_config()
            # Don't send password back to frontend for security
            safe_config = {
                "enabled": config["enabled"],
                "email": config["credentials"]["email"],
                "persistent_session": config["persistent_session"],
                "has_password": bool(config["credentials"]["password"])
            }
            return jsonify({"config": safe_config, "success": True})
        except Exception as e:
            logger.error(f"❌ Error getting auth config: {str(e)}")
            return jsonify({"error": "Failed to get auth config", "success": False}), 500

    # ---------------- Route: Update authentication config ----------------
    @app.route("/auth-config", methods=["POST"])
    def update_auth_config():
        """Update authentication configuration"""
        try:
            data = request.get_json()
            if not data:
                return jsonify({"error": "No data provided", "success": False}), 400

            current_config = load_auth_config()
            
            # Update configuration
            if "enabled" in data:
                current_config["enabled"] = bool(data["enabled"])
            
            if "email" in data:
                current_config["credentials"]["email"] = str(data["email"]).strip()
            
            if "password" in data:
                current_config["credentials"]["password"] = str(data["password"])
            
            if "persistent_session" in data:
                current_config["persistent_session"] = bool(data["persistent_session"])

            # Save updated config
            if save_auth_config(current_config):
                logger.info("✅ Auth config updated successfully")
                return jsonify({
                    "message": "Authentication config updated successfully",
                    "success": True
                })
            else:
                return jsonify({"error": "Failed to save config", "success": False}), 500

        except Exception as e:
            logger.error(f"❌ Error updating auth config: {str(e)}")
            return jsonify({"error": "Failed to update auth config", "success": False}), 500

    # ---------------- Route: Test authentication ----------------
    @app.route("/test-auth", methods=["POST"])
    def test_auth():
        """Test Amazon authentication with current credentials"""
        async def run():
            browser = None
            playwright = None
            try:
                auth_config = load_auth_config()
                if not auth_config['enabled']:
                    return {"error": "Authentication is disabled", "success": False}
                
                if not auth_config['credentials']['email'] or not auth_config['credentials']['password']:
                    return {"error": "Email and password must be configured", "success": False}

                browser, playwright, page = await get_browser(headless=True, auto_login=True)
                
                # Check if we're logged in by looking for account info
                await page.goto("https://www.amazon.com", timeout=15000)  # Reduced from 30s to 15s
                await page.wait_for_load_state("domcontentloaded")
                await human_delay(1000, 1500)  # Reduced from 2-3s to 1-1.5s

                try:
                    account_indicator = page.locator('#nav-link-accountList')
                    if await account_indicator.count() > 0:
                        account_text = await account_indicator.text_content()
                        if account_text and 'Hello' in account_text and 'sign in' not in account_text.lower():
                            return {
                                "message": "Authentication successful", 
                                "account_info": account_text.strip(),
                                "success": True
                            }
                except:
                    pass

                return {"error": "Authentication failed - could not verify login", "success": False}

            except Exception as e:
                logger.error(f"❌ Error testing auth: {str(e)}")
                return {"error": f"Test failed: {str(e)}", "success": False}
            finally:
                if browser:
                    await browser.close()
                if playwright:
                    await playwright.stop()

        loop = asyncio.new_event_loop()
        asyncio.set_event_loop(loop)
        return jsonify(loop.run_until_complete(run()))

    # ---------------- Route: Authentication status ----------------
    @app.route("/auth-status", methods=["GET"])
    def auth_status():
        """Return current authentication status without attempting login"""
        async def run():
            browser = None
            playwright = None
            try:
                config = load_auth_config()
                enabled = bool(config.get('enabled'))

                # If disabled, report immediately
                if not enabled:
                    return {
                        "enabled": False,
                        "is_logged_in": False,
                        "email": "",
                        "account_info": "",
                        "success": True
                    }

                # Start browser, restore cookies if present, but do NOT attempt login
                browser, playwright, page = await get_browser(headless=True, auto_login=False)

                # Quick validation by checking account element
                await page.goto("https://www.amazon.com", timeout=15000)
                await page.wait_for_load_state("domcontentloaded")
                await human_delay(300, 600)

                is_logged_in = False
                account_info = ""
                try:
                    account_indicator = page.locator('#nav-link-accountList')
                    if await account_indicator.count() > 0:
                        account_text = await account_indicator.text_content()
                        if account_text and 'Hello' in account_text and 'sign in' not in account_text.lower():
                            is_logged_in = True
                            account_info = (account_text or '').strip()
                except Exception:
                    pass

                return {
                    "enabled": True,
                    "is_logged_in": is_logged_in,
                    "email": config.get('credentials', {}).get('email', ''),
                    "account_info": account_info,
                    "success": True
                }
            except Exception as e:
                logger.error(f"❌ Error checking auth status: {str(e)}")
                return {
                    "enabled": False,
                    "is_logged_in": False,
                    "email": "",
                    "account_info": "",
                    "error": f"Status check failed: {str(e)}",
                    "success": False
                }
            finally:
                try:
                    if browser:
                        await browser.close()
                    if playwright:
                        await playwright.stop()
                except Exception:
                    pass

        loop = asyncio.new_event_loop()
        asyncio.set_event_loop(loop)
        return jsonify(loop.run_until_complete(run()))

    # ---------------- Route: Clear authentication data ----------------
    @app.route("/clear-auth", methods=["POST"])
    def clear_auth():
        """Clear stored authentication data and session cookies"""
        try:
            # Reset config to defaults
            default_config = AUTO_AUTH_CONFIG.copy()
            save_auth_config(default_config)
            
            # Remove session file if it exists
            session_file = os.path.join(os.path.dirname(__file__), default_config['session_file'])
            if os.path.exists(session_file):
                os.remove(session_file)
                logger.info("✅ Session file removed")
            
            return jsonify({
                "message": "Authentication data cleared successfully",
                "success": True
            })
        except Exception as e:
            logger.error(f"❌ Error clearing auth data: {str(e)}")
            return jsonify({"error": "Failed to clear auth data", "success": False}), 500

    # Serve frontend files
    @app.route('/')
    def serve_frontend():
        return send_file('../index.html')

    @app.route('/<path:filename>')
    def serve_static(filename):
        return send_file(f'../{filename}')

    return app

app = create_app()
CORS(app)

if __name__ == "__main__":
    app.run(port=5000, debug=True)
